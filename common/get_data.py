import os
import openpyxl


def get_login_data():
    path = os.path.dirname(os.path.abspath(__file__))  # 获取当前脚本所在的绝对路径
    filename = os.path.join(path, '..', 'data', 'logindata.xlsx')  # 构建 logindata.xlsx 文件的路径
    workbook = openpyxl.load_workbook(filename)  # openpyxl打开xlsx文件
    sheet = workbook.active  # 获取活动工作表
    lis = []
    for row in sheet.iter_rows(min_row=2):
        row_data = []
        for cell in row:  # 遍历当前行中的单元格
            row_data.append(cell.value)  # 将当前单元格的值添加到当前行的数据列表中
        lis.append(row_data)
    return lis


# def data():
#     def get_login_data_from_excel(file_path, sheet_name):
#         login_data = []
#         workbook = openpyxl.load_workbook(file_path)
#         sheet = workbook[sheet_name]
#
#         for row in sheet.iter_rows(min_row=2):
#             row_data = []
#             for cell in row:  # 遍历当前行中的单元格
#                 row_data.append(cell.value)  # 将当前单元格的值添加到当前行的数据列表中
#             login_data.append(row_data)
#         return login_data
# def demo():
#     file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data', 'logindata.xlsx')
#     sheet_name = 'Sheet1'
#     login_data = get_login_data_from_excel(file_path, sheet_name)
#
#     sheet_name = 'login'
#     path = os.path.dirname(os.path.abspath(__file__))
#     file_path= os.path.join(path,'..')


if __name__ == '__main__':
    testcase = get_login_data()
    print(testcase)
